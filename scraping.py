from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import time

# Setup driver
options = Options()
options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36')
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# List ticker crypto
tickers = ["bitcoin", "ethereum", "bnb"]

results = []

for ticker in tickers:
    url = f"https://coinmarketcap.com/currencies/{ticker}"
    driver.get(url)
    time.sleep(3)  # Tunggu render
    
    try:
        price = driver.find_element(By.XPATH, '//span[contains(@data-test, "text-cdp-price-display")]').text
        change = driver.find_element(By.XPATH, '//p[contains(@class, "change-text")]').text

        Parent = driver.find_elements(By.XPATH, '//div[contains(@data-role, "group-item")]')

        # Market Cap
        marketCapChild = Parent[0].find_element(By.XPATH, './/div[contains(@class, "CoinMetrics_overflow-content__tlFu7")]')
        marketCap = marketCapChild.text

        # Volume 24h
        volumeChild = Parent[1].find_element(By.XPATH, './/div[contains(@class, "CoinMetrics_overflow-content__tlFu7")]')
        volume = volumeChild.text

        # Circulation Supply
        circulatingSupplyChild = Parent[6].find_element(By.XPATH, './/div[contains(@class, "CoinMetrics_overflow-content__tlFu7")]')
        circulatingSupply = circulatingSupplyChild.text

        # Volume Market Cap Ratio
        volumeMarketCapRatioFormula = volume / marketCap

        if volumeMarketCapRatioFormula > 0.1:
            volumeMarketCapRatio = "Market lagi Hype"
        elif 0.05 > volumeMarketCapRatioFormula < 0.1:
            volumeMarketCapRatio = "Market lagi Konsolidasi"
        else:
            volumeMarketCapRatio = "Market lagi Lesu"

        #

        # Bersihkan change
        change_value = float(change.replace('% (1d)', '').strip())

        # Tentukan keputusan
        if change_value > 1:
            decision = "Buy"
        elif change_value < -1:
            decision = "Sell"
        else:
            decision = "Hold"

        results.append({
            "Ticker": ticker,
            "Price": price, 
            "Change (1d)": change.replace('(1d)', ''), 
            "Market Cap": marketCap,
            "Volume (24h)": volume,
            "Circulating Supply": circulatingSupply,
            "Market Hype": volumeMarketCapRatio,
            "Decision": decision
        })

    except Exception as e:
        print(f"Data gagal diambil untuk {ticker}: {e}")

driver.quit()

# Convert ke DataFrame
df = pd.DataFrame(results)
df.index += 1  # Biar No nya mulai dari 1 (optional)

# Create Excel workbook langsung
wb = Workbook()
ws = wb.active
ws.title = "Crypto Data"

# Styling
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

start_row = 2
start_col = 2

# Masukin header manual: No + Kolom2
headers = ["No"] + df.columns.tolist()

for c_idx, header in enumerate(headers):
    cell = ws.cell(row=start_row, column=start_col + c_idx, value=header)
    cell.font = bold_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Masukin data
for r_idx, row in enumerate(df.itertuples(), start=1):
    row_data = [r_idx] + list(row[1:])  # r_idx itu No
    for c_idx, value in enumerate(row_data):
        cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

# Save file
wb.save('crypto_dataset_with_decision_styled.xlsx')

print(df)